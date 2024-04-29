import openpyxl as xl
from openpyxl.styles import Font

#create a new excel document
wb = xl.Workbook()

ws = wb.active

ws.title = 'First Sheet'

wb.create_sheet(index=1,title='Second Sheet')

wb.save("PythontoExcel.xlsx")

ws['A1'] = 'Invoice'

ws['A1'].font = Font(name="Times New Roman",size=24,bold=True)

headerfont = Font(name='TImes New Roman',size=24,bold=True)

ws['A2'] = 'Tires'
ws['A3'] = 'Brakes'
ws['A4'] = 'Allignment'

ws.merge_cells('A1:B1')

ws['B2'] = 450
ws['B3'] = 225
ws['B4'] = 150

#unmerge cells
#ws.unmerge_cells['A1':B1']

ws['A8'] = 'Total'
ws['A8'].font = Font(size=16,bold=True)

ws['B8'] = '=Sum(B2:B4)'

ws.column_dimensions['A'].width = 25

wb.save("PythontoExcel.xlsx")

write_sheet = wb['Second Sheet']
read_wb = xl.load_workbook('ProduceReport.xlsx')
read_ws = read_wb['ProduceReport']

for i in range(1,read_ws.max_row+1):
    write_sheet['A'+str(i)] = read_ws.cell(i,1).value
    write_sheet['B'+str(i)] = read_ws.cell(i,2).value
    write_sheet['C'+str(i)] = read_ws.cell(i,3).value
    write_sheet['D'+str(i)] = read_ws.cell(i,4).value


write_sheet['A43'] = 'Grand Total'
write_sheet['A43'].font = Font(size=16,bold=True)

write_sheet['C43'] = '=Sum(C2:C41)'
write_sheet['D43'] = '=Sum(D2:D41)'


write_sheet['A44'] = 'Avg'
write_sheet['A44'].font = Font(size=16,bold=True)

write_sheet['C44'] = '=AVERAGE(C2:C41)'
write_sheet['D44'] = '=AVERAGE(D2:D41)'

for cell in write_sheet['C:C']:
    cell.number_format = '#,##0'

for cell in write_sheet['D:D']:
    cell.number_format = u'"$ "#,##0.00'


write_sheet.column_dimensions['A'].width = 25
wb.save("PythontoExcel.xlsx")